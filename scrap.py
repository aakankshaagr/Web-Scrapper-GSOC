import time
import openpyxl
from selenium import webdriver
from bs4 import BeautifulSoup
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC

# initialize web driver
driver = webdriver.Chrome()
past_orgs=[]   

# to automate scrapping on year 2021-2016
for year in range(2021,2015,-1):
    
    #url for each year
    url="https://summerofcode.withgoogle.com/archive/"+str(year)+"/organizations"
    driver.get(url)
    
    #wait till DOM is loaded
    try:
        element = WebDriverWait(driver, 10).until(
            EC.presence_of_element_located((By.XPATH, "/html/body/app-root/app-layout/mat-sidenav-container/mat-sidenav-content/div/main/app-organizations/app-orgs-grid/section/div/div"))
        )
    except:
        print("website not found")

    #extract page source
    element= driver.page_source
    
    #convert it to html
    soup = BeautifulSoup(element, "html.parser")
    
    #extract card innerHTML
    result=soup.find_all("div",class_="card")
    #extract required information 
    for element in result:
        count={}
        # extract org name
        name=element.find("div",class_="name")
        org_name=name.text.strip()
        # extract org description
        desc=element.find("div",class_="short-description")
        org_desc=desc.text
        # extract org link
        link=element.find("a", {"class": "content"}).attrs['href'].strip()
        org_link="https://summerofcode.withgoogle.com"+str(link)
        # add new org in list and update year if org is repeated
        for i in range(len(past_orgs)):
            
            if org_name == past_orgs[i]["Name"]:
                #updating repeated orgs data
                past_orgs[i][year]="yes"
                past_orgs[i]["rep"]+=1
                break
                
        else:
            #adding new orgs
            count["Name"]=org_name
            count["link"]=org_link
            count["description"]=org_desc
            for yr in range(2021,2015,-1):
                if yr==year:
                    count[yr]="yes"
                else:
                    count[yr]="no"
            count["rep"]=1
            past_orgs.append(count)

# adding extracted past orgs to excel sheet
#opening excel sheet
wb = openpyxl.Workbook()
#getting current active sheet
ws=wb.active
# setting name of sheet
ws.title='GSOC'
# setting headers of each row
ws['A1'] = 'Name'
ws['B1'] = 'Link'
ws['C1']='Description'
ws['D1']="2021"
ws['E1']="2020"
ws['F1']="2019"
ws['G1']="2018"
ws['H1']="2017"
ws['I1']="2016"
ws['J1']="Count"
# inserting org data
i=2
for orgs in past_orgs:
    
    ws["A"+str(i)]=orgs["Name"]
    ws["B"+str(i)]=orgs["link"]
    ws["C"+str(i)]=orgs["description"]
    ws["D"+str(i)]=orgs[2021]
    ws["E"+str(i)]=orgs[2020]
    ws["F"+str(i)]=orgs[2019]
    ws["G"+str(i)]=orgs[2018]
    ws["H"+str(i)]=orgs[2017]
    ws["I"+str(i)]=orgs[2016]
    ws["J"+str(i)]=orgs["rep"]
    i+=1
#saving excel sheet
wb.save('Gsoc.xlsx')
